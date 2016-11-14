import openpyxl
from operator import attrgetter
import datetime
from copy import deepcopy

#this class is a MoneyLover row
#each row in the excel file becomes a class
#we use a list of these to represent the whole file
class mlRow():
	#this one tuple passed as arg thing helped with import, but its a bane for adding entries...
	def __init__(self, wbRow):
		#wbRow is the row list as passed from the original sheet.rows[index] list returned
		#this is about to get really ugly and hackish T_T
		#print type(wbRow)
		#for x in wbRow: print x
		#this first bit is for if we pass a Cell Sheet object tuple
		try:
			try: self.id = int(wbRow[0].value)
			except ValueError: self.id = wbRow[0].value
			self.category = wbRow[1].value
			self.amount = wbRow[2].value
			self.note = wbRow[3].value
			self.wallet = wbRow[4].value
			self.currency = wbRow[5].value
			self.date = wbRow[6].value
		#and this is if we pass straight data.. e.g. there is no .value attribute
		except AttributeError:
			self.id = wbRow[0]
			self.category = wbRow[1]
			self.amount = wbRow[2]
			self.note = wbRow[3]
			self.wallet = wbRow[4]
			self.currency = wbRow[5]
			self.date = wbRow[6]
		#yeah.. I know... if you have a better idea, contribute to the project
		#don't hate, levitate

	def display(self):
		print "ID: " + str(self.id)
		print "Category: " + str(self.category)
		print "Amount: $" + str(self.amount)
		print "Note: " + str(self.note)
		print "Wallet: " + str(self.wallet)
		print "Currency: " + str(self.currency)
		print "Date: " + str(self.date)
		print ""

	#returns true is self and row contain the same data
	def compare(self, row):
		if(self.id == row.id and
			self.category == row.category and
			self.amount == row.amount and
			self.note == row.note and
			self.wallet == row.wallet and
			self.currency == row.currency and
			self.date == row.date):
			return True
		else:
			return False

#take a file name as argument and return tuple containing previously mentioned list of mlRow classes and header
def loadMLWorkbook(workbkfname):
	#load MoneyLover generated workbook
	originalWorkbook = openpyxl.load_workbook(workbkfname)
	originalSheet = originalWorkbook.active

	#list of rows in mlRow class format
	rowClassList = []
	header = mlRow(tuple(originalSheet.rows)[0])		#figure out how to keep green color for header

	for line in range(1, originalSheet.max_row):		#start at 1, line 0 is category names (header)
		#print type((originalSheet.rows))
		rowClassList.append(mlRow(tuple(originalSheet.rows)[line]))

	#convert all dates to datetime objects
	#we may want to put this in the mlRow constructor
	for elem in rowClassList:
                elem.date = datetime.datetime.strptime(elem.date, "%m/%d/%Y")

	return rowClassList, header	#returns list of mlRow objects, header


#take the mlRow class list and generate a double indexed array
#return said double indexed array
#the purpose for this is that it is easier to work with in some contexts
#for example, writing out to a new excel file
def generateArray(rowClassList):	#takes argument of mlRows class and converts to a conventional array
	numRows = len(rowClassList)
	rowArrayList = []		#array is [rows][columns]
	
	for i in range(0, numRows):
		rowArrayList.append([])		#create first index of array

		#create second index
		rowArrayList[i].append(rowClassList[i].id)
		rowArrayList[i].append(rowClassList[i].category)
		rowArrayList[i].append(rowClassList[i].amount)
		rowArrayList[i].append(rowClassList[i].note)
		rowArrayList[i].append(rowClassList[i].wallet)
		rowArrayList[i].append(rowClassList[i].currency)
		rowArrayList[i].append(rowClassList[i].date)

	return rowArrayList

def addEntry(workbklist, category, amount, note, wallet, currency, date):
	if(type(date) is str): 
		date = convertStrToDatetime(date)	#allows for passing either string or datetime obj to function
	lastID = sortByID(workbklist)[-1].id	#sort by id to get last id
	lst = [lastID+1, category, amount, note, wallet, currency, date]	#load up a list to pass to constructor
	workbklist.append(mlRow(lst))		#construct
	return workbklist

#remove entry from workbklist by object address
#the passed list is modified, and then returned as an extra feature
def removeEntry(workbklist, obj):
	workbklist.remove(obj)
	return workbklist

#may need to create a getEntryByID() function to have passable objects
# ...copying and pasting code is bad >_<
#here we have a binary search to find an object by it's unique ID int
#returns the object, -1 on error
#getByID() is wrapper for idSearch() so we can sort the list
def getByID(workbklist, id):		#TEST
	idSortedList = sortByID(workbklist)
	return idSearch(idSortedList, id)

#this is the pure binary search function, it will not be documented in the documentation
#use by calling getByID(), returns object by unique ID, list must be id sorted already
def idSearch(idSortedList, id):		#TEST
	#base case
        if len(idSortedList) == 1:
                if(idSortedList[0].id != id): return -1
                return idSortedList[0]

        mid = len(idSortedList)/2
        #print "length: " + str(len(lst)) + " mid: " + str(mid)
        if idSortedList[mid].id <= id:
                return getByID(idSortedList[mid:], id)
        elif idSortedList[mid].id > id:
                return getByID(idSortedList[:mid], id)

#take the mlRows list as an argument, sort it, and return the sorted list
def sortByCategory(workbklist):	

	return sorted(workbklist, key=attrgetter('category'))

def sortByDate(workbklist):

	return sorted(workbklist, key=attrgetter('date'))

def sortByID(workbklist):

	return sorted(workbklist, key=attrgetter('id'))

#splice mlRows list based on criteria, return list of uniform category lists
def spliceByCategory(workbklist):
	if len(workbklist) == 0: return []
	catsortedlist = sortByCategory(workbklist)
	workinglist = []
	workinglist.append(catsortedlist[0])		#what if passed empty list?
	toreturn = []
	i = 0
	while(i < len(catsortedlist)-1):
		if(catsortedlist[i].category == catsortedlist[i+1].category):
			workinglist.append(catsortedlist[i+1])
		else:
			toreturn.append(workinglist)
			workinglist = []
			workinglist.append(catsortedlist[i+1])
		i+=1

	toreturn.append(workinglist)
	return toreturn

#takes three arguments, the first two are date strings in mm/dd/yyyy format, the third is the list to splice from
#returns the spliced list sorted by date, returns -1 if unable to splice
#returns empty list if no items in range
def spliceDateRange(startdate, enddate, workbklist):
	if len(workbklist) == 0: return []
	try:
		if(type(startdate) is str): dtstartdate = convertStrToDatetime(startdate)
		else: dtstartdate = startdate
		if(type(enddate) is str): dtenddate = convertStrToDatetime(enddate)
		else: dtenddate = enddate
		#dtstartdate = datetime.datetime.strptime(startdate, "%m/%d/%Y")
		#dtenddate = datetime.datetime.strptime(enddate, "%m/%d/%Y")
	except:
		print "incorrect date format passed"
		return -1

	if(dtstartdate > dtenddate):	#if startdate is after enddate
		return -1		#return -1 for error
	

	workinglist = []

	#there may be a chance that one of the next three statements is removing a row or two
	#correct entries are returned but the indexes might be wrong...
	#almost definitely not but keeping comment just in case
	datesortedlist = sortByDate(workbklist)		#sort the list by date

	#if we are only looking for one date
	if(dtstartdate == dtenddate):
		retIndexes = dateSearchAll(startdate, datesortedlist)
		if(retIndexes == -1): return -1					#if it's not in the list
		for i in retIndexes: workinglist.append(datesortedlist[i])
		return workinglist

	#check if start and end dates are in the list, if not find the closest date in list within the range and go with that
	if (dateSearch(dtstartdate, datesortedlist) == -1):
		dtstartdate = dateSearchClosest(dtstartdate, 1, datesortedlist)
	if (dateSearch(dtenddate, datesortedlist) == -1):
		dtenddate = dateSearchClosest(dtenddate, -1, datesortedlist)

	if(dtenddate == -1 or dtstartdate == -1): return workinglist	#if nothing in bounds, return empty list

	startindexes = dateSearchAll(dtstartdate, datesortedlist)
	endindexes = dateSearchAll(dtenddate, datesortedlist)
	#for x in startindexes:
	#	datesortedlist[x].display()
	#for x in endindexes:
	#	datesortedlist[x].display()

	begin = startindexes[0]
	end = endindexes[-1]
	#print endindexes
	#print endindexes[-1]
	for i in range(begin, end+1):
		workinglist.append(datesortedlist[i])

	return workinglist

def convertStrToDatetime(str):
	return datetime.datetime.strptime(str, "%m/%d/%Y")

def convertDatetimeToStr(dt):
	return datetime.datetime.strftime(dt, "%m/%d/%Y")

#MUST IMPLEMENT CHECKS THAT LISTS PASSED HAVE ENTRIES

#date is date string
#directionstep is a positive or negative int
#workbklist is a date sorted list
#returns next closest date in direction
def dateSearchClosest(date, directionstep, workbklist):
	if(type(date) is str): date = convertStrToDatetime(date)

	#check for bounds
	if(directionstep == 0):
		print "direction step must not be 0!"
		return -2
	elif(date < workbklist[0].date and directionstep < 0):
		return -1
	elif(date > workbklist[len(workbklist)-1].date and directionstep > 0):
		return -1

	while(dateSearch(date, workbklist) == -1):
		date += datetime.timedelta(days=directionstep)	#to increment or decrement date 
	return date

#recursive function for binary search
#returns object of last instance found of given date or -1 if not found
#takes a datetime object and a sorted mlRow class list as arguments
#for internal library use, reference directly from your code at your own peril
def dateSearch(date, lst):
	#base case here
	if len(lst) == 1:
		if(lst[0].date != date): return -1
		return lst[0]

	mid = len(lst)/2
	#print "length: " + str(len(lst)) + " mid: " + str(mid)
	if lst[mid].date <= date:
		return dateSearch(date, lst[mid:])
	elif lst[mid].date > date:
		return dateSearch(date, lst[:mid])

#we must implement a way to enforce sorted vs unsorted lists

#takes a date string in "mm/dd/yyyy" format and a sorted mlRow list as arguments
#returns a list of the indexes of all entries on that date
def dateSearchAll(date, sortedlist):
	try:
		if(type(date) is str): date = convertStrToDatetime(date)
		#date = datetime.datetime.strptime(date, "%m/%d/%Y")		#convert date to a datetime object
	except:
		print "invalid format"
		return -1

	#get the index from the object returned by dateSearch()
	obj = dateSearch(date, sortedlist)
	#print "IN DATESEARCHALL()"
	#print date
	#obj.display()
	if(obj == -1):			#check for if date not in sortedlist
		return -1
	index = sortedlist.index(obj)
	workinglist = []
	i = index
	
	#scan list forward
	#this bit might be unnecessary or inefficient and modifiable, more investigation needed
		#seems to go through first loop one time at beginning
	while(i < len(sortedlist) and sortedlist[index].date == sortedlist[i].date):
		workinglist.append(i)
		i+=1
		#print "TESTING"
	#scan list backward
	i = index-1
	while(i >= 0 and sortedlist[index].date == sortedlist[i].date):
		workinglist.append(i)
		i-=1
		#print "TESTING2"
	workinglist.sort()
	return workinglist

#provide mlRows list as an argument and write it out to a new excel file specified by string newFileName
def exportToNewWorkbook(header, mlWkbkClassList, newFileName):
	#print mlWkbkClassList
	toWrite = deepcopy(mlWkbkClassList)		#make copy so we don't modify the original list
	#print toWrite
	for elem in toWrite:
		elem.date = datetime.datetime.strftime(elem.date, "%m/%d/%Y")   #restore date to original text format

	#we use the generateArray() function so that we can double iterate
	#with for loops and easily assign the cells in the new sheet
	toWrite.insert(0, header)
	genArray = generateArray(toWrite)

	newWorkbook = openpyxl.Workbook()
	newSheet = newWorkbook.active

	i = 1		#sometimes I wish I was working in C
	for arrObj in genArray:		#for each row
		for j in range(0,7):	#for each column
			newSheet.cell(row=i, column=(j+1)).value = arrObj[j]
		i+=1

	newWorkbook.save(newFileName)
